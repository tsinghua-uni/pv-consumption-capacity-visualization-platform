import random
import schedule
import time
import pandas as pd
from pyecharts import options as opts
from pyecharts.charts import Map
from pyecharts.globals import ThemeType
from pyecharts.charts import PictorialBar
from openpyxl import load_workbook
from pyecharts.globals import SymbolType

# Function to generate random data and update the excel file
def update_excel():
    data = load_workbook('data_cities-dynamic.xlsx')
    table = data.active
    for i in range(2, 3801):  # B2 to B3800
        table['B' + str(i)] = random.randint(100, 10000)  # generate random number between 1 and 10000
    data.save('data_cities-dynamic.xlsx')
def update_excel2():
    data = load_workbook('data_cities-dynamic2.xlsx')
    table = data.active
    for i in range(2, 3801):  # B2 to B3800
        table['B' + str(i)] = random.randint(100, 10000)  # generate random number between 1 and 10000
    data.save('data_cities-dynamic2.xlsx')
def generate_map1():
    # Read the Excel file
    df = pd.read_excel('data_cities-dynamic.xlsx')

    # Convert the DataFrame to a list of tuples
    data = list(df.itertuples(index=False, name=None))

    c = (
        Map(init_opts=opts.InitOpts(page_title="PV Consumption Capacity Visualization Platform",bg_color="#FFFAFA",width="800px", height="600px", theme=ThemeType.WHITE))
# Map(init_opts=opts.InitOpts(bg_color="#FFFAFA", theme=ThemeType.ESSOS))
        .add(
            "Show/Hide",
            data,
            "china-cities",
            label_opts=opts.LabelOpts(is_show=False),
            is_map_symbol_show=False,
            
        )
        # # 设置坐标属性，显示省份名
        # .set_series_opts(
        # label_opts=opts.LabelOpts(is_show=True))
        .set_global_opts(
            title_opts=opts.TitleOpts(title="PV Consumption Capacity", subtitle="Visualization Platform"),
            visualmap_opts=opts.VisualMapOpts(
                max_=10000,
                range_color=["#FF0000", "#FFA500", "#FFFF00", "#27cc7f", "#eeeeee"],
                is_piecewise=True,
                pos_top='50%',  # Adjust this value
                pos_left='5%',  # Adjust this value
                pieces=[
                    {"min": 1, "max": 2000, "color": "#FF0000", "label": "Low consumption capacity"},
                    {"min": 201, "max": 4000, "color": "#FFA500", "label": "Medium-low consumption capacity"},
                    {"min": 401, "max": 6000, "color": "#FFFF00", "label": "Medium-high consumption capacity"},
                    {"min": 601, "max": 10000, "color": "#27cc7f", "label": "High consumption capacity"},
                    # No data，grey
                    {"min": -1000, "max": 0, "color": "#eeeeee", "label": "No data"}
                ]
            ),
        )
    )
    c.render("map1.html")  # render map1 to "map1.html"

def generate_map2():
    # Read the Excel file,Read C column
    df = pd.read_excel('data_cities-dynamic2.xlsx')

    # Convert the DataFrame to a list of tuples
    data = list(df.itertuples(index=False, name=None))

    c = (
        Map(init_opts=opts.InitOpts(page_title="PV Consumption Capacity Visualization Platform",bg_color="#FFFAFA",width="800px", height="600px", theme=ThemeType.WONDERLAND))
# Map(init_opts=opts.InitOpts(bg_color="#\\", theme=ThemeType.ESSOS))
        .add(
            "Show/Hide",
            data,
            "china-cities",
            label_opts=opts.LabelOpts(is_show=False),
            is_map_symbol_show=False,
            
        )
        # # 设置坐标属性，显示省份名
        # .set_series_opts(
        # label_opts=opts.LabelOpts(is_show=True))
        .set_global_opts(
            title_opts=opts.TitleOpts(title="Carbon Footprint", subtitle="City Version"),
            visualmap_opts=opts.VisualMapOpts(
                max_=10000,
                range_color=["#FF0000", "#FFA500", "#FFFF00", "#27cc7f", "#eeeeee"],
                is_piecewise=True,
                pos_top='50%',  # Adjust this value
                pos_left='5%',  # Adjust this value
                pieces=[
                    {"min": 1, "max": 2000, "color": "#FF0000", "label": "High Carbon Footprint"},
                    {"min": 201, "max": 4000, "color": "#FFA500", "label": "Medium-high Carbon Footprint"},
                    {"min": 401, "max": 6000, "color": "#FFFF00", "label": "Medium-low Carbon Footprint"},
                    {"min": 601, "max": 10000, "color": "#27cc7f", "label": "Low Carbon Footprint"},
                    # No data，grey
                    {"min": -1000, "max": 0, "color": "#eeeeee", "label": "No data"}
                ]
            ),
        )
    )
    c.render("map2.html")  # render map2 to "map2.html"


def generate_maps():
    generate_map1()
    generate_map2()

# Create a new HTML file that includes both maps using iframes
with open("maps.html", "w") as f:
    f.write("""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {
                display: flex;
                flex-direction: column;
                align-items: center;
                justify-content: center;
                width: 100vw;
                height: 100vh;
                margin: 0;
                padding: 0;
                overflow: hidden;
                font-family: Arial, sans-serif;
            }
            h1 {
                margin-bottom: 20px;
            }
            #map-container {
                display: flex;
                flex-direction: row;
                width: 100vw;
                height: calc(100vh - 60px);
            }
            iframe {
                width: 50vw;
                height: 100%;
                border: none;
                overflow: hidden;
            }
            #controls {
                display: flex;
                justify-content: space-between;
                width: 300px;
                margin-bottom: 20px;
            }
            button {
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                color: white;
                cursor: pointer;
            }
            .start {
                background-color: #27cc7f;
            }
            .stop {
                background-color: #cc2727;
            }
            .refresh {
                background-color: #277fcc;
            }
        </style>
    <script>
    var autoReload = null;

    function startAutoReload() {
        autoReload = setInterval(function(){
            window.location.reload(1);
        }, 2000); // Reload every 2000ms
    }

    function stopAutoReload() {
        if (autoReload) {
            clearInterval(autoReload);
            autoReload = null;
            alert("Auto reload stopped.");  // Show message
        }
    }

    function refreshPage() {
        location.reload();
    }

    // Call startAutoReload immediately after defining it
    startAutoReload();
    </script>
    </head>
    <body>
        <h1>PV Consumption Capacity Visualization Platform</h1>
        <div id="clock">
        <iframe width="80" height="20" src="https://vclock.com/embed/time/#theme=0&ampm=1&showdate=1" frameborder="0" allowfullscreen></iframe>
            </div>
        <div id="controls">
            <button class="start" onclick="startAutoReload()">Start Auto Reload</button>
            <button class="stop" onclick="stopAutoReload()">Stop Auto Reload</button>
            <button class="refresh" onclick="refreshPage()">Refresh</button>
        </div>
        <div id="map-container">
            <iframe src="map1.html"></iframe>
            <iframe src="map2.html"></iframe>
        </div>
    </body>
    </html>
    """)
# Function to update excel and generate map
def job():
    update_excel()
    update_excel2()
    generate_maps()

# Schedule job every 5 second
schedule.every(3).seconds.do(job)

# Keep running the script
while True:
    schedule.run_pending()
    time.sleep(2)